import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter


class XlFormat:
    def __init__(self):
        self.FONT_NAME = "游ゴシック"

    def formatting(self, sheet):

        # header設定
        row_length = len(list(sheet.rows)[0])

        fill = PatternFill(patternType="solid", fgColor="d3d3d3")
        for i in range(row_length):
            sheet.cell(row=1, column=i + 1).fill = fill

        # 全セルの書式設定
        side = Side(style="thin", color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        alignment = Alignment(wrapText=True, horizontal="left", vertical="top")
        font = Font(name=self.FONT_NAME)

        for col in sheet.columns:
            max_length_in_col = 0
            column = col[0].column

            for cell in col:
                cell.border = border
                cell.alignment = alignment
                cell.font = font

                # cell.value を改行コードで分割し、col内の最大長文字幅を取得
                values = str(cell.value).splitlines()
                max_length_in_cell = 0
                for value in values:
                    if len(str(value)) > max_length_in_cell:
                        max_length_in_cell = len(str(value))

                if max_length_in_cell > max_length_in_col:
                    max_length_in_col = max_length_in_cell

            # 列幅調整
            col_width = (max_length_in_col + 5) * 1.2
            sheet.column_dimensions[get_column_letter(column)].width = col_width
